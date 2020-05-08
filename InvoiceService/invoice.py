import os
import shutil
from datetime import date
from datetime import datetime

import openpyxl
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

pdfmetrics.registerFont(TTFont('HelveticaNeue', 'HelveticaNeueCyr-Light.ttf'))

today = date.today()

#make a copy of the invoice to work with
src_dir="blank_invoice.pdf"
dst_dir="Invoice.pdf"
shutil.copy(src_dir,dst_dir)

#import the sheets from the excel file
try:
    clientTable = openpyxl.load_workbook('client_info.xlsx')
    clientSheet = clientTable['clients']
except:
    print("Could not access the Client Table! Shutting down!")
    raise SystemExit
try:
    productTable = openpyxl.load_workbook('product_info.xlsx')
    productSheet = productTable['products']
except:
    print("Could not access the Product Table! Shutting down!")
    raise SystemExit

#Page information
page_width = 2156
page_height = 3050

class CompanyClient:
    CUI = ''
    Inreg = ''


def checkIfCompany(field):
    if(field == "x"):
        return True
    else:
        return False

# get customer and product data
customer = clientSheet.cell(row = 2, column = 1).value
print(customer, end='  ')
if(checkIfCompany(clientSheet.cell(row = 2, column = 2).value)):
    company = CompanyClient()
    company.CUI = clientSheet.cell(row = 2, column = 3).value
    company.Inreg = clientSheet.cell(row = 2, column = 4).value
    print(company.CUI, company.Inreg, end='  ')
address = clientSheet.cell(row = 2, column = 5).value
print(address)

productId = productSheet.cell(row = 2, column = 1).value
productName = productSheet.cell(row = 2, column = 2).value
productPrice = productSheet.cell(row = 2, column = 3).value

print(productId, ' ', productName, ' ', productPrice)

#g get invoice series
invoiceSeriesFile = open("invoice_series.txt","r")
invoiceSeries = invoiceSeriesFile.readline()
invoiceSeriesFile.close()
print("Invoice series: ", invoiceSeries)

# get invoice number
invoiceNumberFile = open("invoice_number.txt", "r")
invoiceNumber = invoiceNumberFile.readline()
invoiceNumberFile.close()
print("Invoice number: ", invoiceNumber)

# update invoice number
invoiceNumberFile = open("invoice_number.txt", "w")
newNr = int(invoiceNumber)
newNr = newNr + 1
invoiceNumberFile.write(str(newNr))
invoiceNumberFile.close()
print("New invoice number: ", newNr)

# create a new PDF with Reportlab
can = canvas.Canvas("TextToAdd.pdf", pagesize=A4)
can.setFont('HelveticaNeue', 11)
serialAndDate = "seria AAA Nr " + invoiceNumber
can.drawString(283, 697, serialAndDate)

dateTimeObj = datetime.now()
timestampStr = dateTimeObj.strftime("%d-%m-%Y")
can.drawString(291, 682, timestampStr)

print('Current Timestamp : ', timestampStr)

can.save()



new_pdf = PdfFileReader('TextToAdd.pdf')
existing_pdf = PdfFileReader("Invoice.pdf")

output = PdfFileWriter()

page = existing_pdf.getPage(0)
page.mergePage(new_pdf.getPage(0))
output.addPage(page)
# finally, write "output" to a real file
outputStream = open("ClientInvoice.pdf", "wb")
output.write(outputStream)
outputStream.close()