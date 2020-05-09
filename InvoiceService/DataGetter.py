import InvoiceMaker
import openpyxl
import DataTypes
    
def checkIfCompany(field):
    if(field == "x"):
        return True
    else:
        return False

#import the sheets from the excel file
try:
    clientTable = openpyxl.load_workbook('client_info.xlsx')
    clientSheet = clientTable['clients']
except:
    print("Could not access the Client Table! Shutting down!")
    raise SystemExit
try:
    productTable = openpyxl.load_workbook('product_info.xlsx')
    productSheet = productTable['products']
except:
    print("Could not access the Product Table! Shutting down!")
    raise SystemExit

orderedProducts = DataTypes.OrderedProducList()
customerInfo = DataTypes.Customer()

# get customer and product data
customerInfo.name = clientSheet.cell(row = 2, column = 1).value
print(customerInfo.name, end='  ')

customerInfo.isCompany = checkIfCompany(clientSheet.cell(row = 2, column = 2).value)

if(customerInfo.isCompany):
    customerInfo.CUI = clientSheet.cell(row = 2, column = 3).value
    customerInfo.nrInreg = clientSheet.cell(row = 2, column = 4).value
    print(customerInfo.CUI, customerInfo.nrInreg)
customerInfo.address = "Adresa: " + clientSheet.cell(row = 2, column = 5).value

# Get IDs and quantity of products to put on invoice

orderedProducts.productId.insert(0, productSheet.cell(row = 2, column = 1).value)
orderedProducts.productName.insert(0, productSheet.cell(row = 2, column = 2).value)
orderedProducts.productPrice.insert(0, productSheet.cell(row = 2, column = 3).value)

orderedProducts.productId.insert(1, productSheet.cell(row = 3, column = 1).value)
orderedProducts.productName.insert(1, productSheet.cell(row = 3, column = 2).value)
orderedProducts.productPrice.insert(1, productSheet.cell(row = 3, column = 3).value)

orderedProducts.productQuantity.insert(0,2)
orderedProducts.productQuantity.insert(1,1)

print(orderedProducts.productId[0], ' ', orderedProducts.productName[0], ' ', orderedProducts.productPrice[0])

#InvoiceMaker.TestMe(orderedProducts)
InvoiceMaker.MakeInvoice(orderedProducts, customerInfo, 55)
