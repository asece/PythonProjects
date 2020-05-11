import InvoiceMaker
import openpyxl
import DataTypes
import Logger

def getInvoiceNumber():
    invoiceNumberFile = open("invoice_number.txt", "r")
    invoiceNumber = invoiceNumberFile.readline()
    invoiceNumberFile.close()
    print("Read invoice number: ", invoiceNumber)
    Logger.Log("{}: {} {}".format(__name__, "Read invoice number: ", invoiceNumber))
    return invoiceNumber

def incrementInvoiceNumber():
    invoiceNumberFile = open("invoice_number.txt", "r")
    invoiceNumber = invoiceNumberFile.readline()
    invoiceNumberFile.close()

    invoiceNumberFile = open("invoice_number.txt", "w")  
    newInvoiceNumber = int(invoiceNumber)
    newInvoiceNumber = newInvoiceNumber + 1
    invoiceNumberFile.write(str(newInvoiceNumber))
    invoiceNumberFile.close()
    print("Wrote new invoice number: ", newInvoiceNumber)
    Logger.Log("{}: {} {}".format(__name__, "Wrote new invoice number: ", newInvoiceNumber))

def getInvoiceSeries():
    invoiceSeriesFile = open("invoice_series.txt","r")
    invoiceSeries = invoiceSeriesFile.readline()
    invoiceSeriesFile.close()
    print("Read invoice series: ", invoiceSeries)
    Logger.Log("{}: {} {}".format(__name__, "Read invoice series: ", invoiceSeries))
    return invoiceSeries

def checkIfCompany(field):
    if(field == "x"):
        return True
    else:
        return False

def getAllClients():
    try:
        clientTable = openpyxl.load_workbook('client_info.xlsx', data_only = True)
        clientSheet = clientTable['clients']
    except:
        print("ERROR: Could not access the Client Table! Shutting down!")
        Logger.Log("{}: {}".format(__name__, "ERROR: Could not access the Client Table! Shutting down!"))
        raise SystemExit

    allCustomerInfo = DataTypes.AllCustomers()

    numberofCustomers = clientSheet.cell(row = 1, column = 10).value
    print("-> Getting information for ", numberofCustomers," customers:")
    Logger.Log("{}: {} {} {}".format(__name__, "-> Getting information for ", numberofCustomers," customers:"))
    for it in range(0,numberofCustomers):
        allCustomerInfo.name.insert(it, clientSheet.cell(row = (2 + it), column = 1).value)
        print(allCustomerInfo.name[it], end='  ')
        Logger.Log("{}: {}".format(__name__, allCustomerInfo.name[it]))
        allCustomerInfo.isCompany.insert(it, checkIfCompany(clientSheet.cell(row = (2 + it), column = 2).value))

        if(allCustomerInfo.isCompany[it]):
            allCustomerInfo.CUI.insert(it, clientSheet.cell(row = (2 + it), column = 3).value)
            allCustomerInfo.nrInreg.insert(it, clientSheet.cell(row = (2 + it), column = 4).value)
            print(allCustomerInfo.CUI[it], allCustomerInfo.nrInreg[it])
            Logger.Log("{}: {} {}".format(__name__, allCustomerInfo.CUI[it], allCustomerInfo.nrInreg[it]))
        allCustomerInfo.address.insert(it, ("Adresa: " + clientSheet.cell(row = (2 + it), column = 5).value))
        print(allCustomerInfo.address[it])
        Logger.Log("{}: {}".format(__name__, allCustomerInfo.address[it]))

    return allCustomerInfo

def getAllProducts():
    try:
        productTable = openpyxl.load_workbook('product_info.xlsx',  data_only = True)
        productSheet = productTable['products']
    except:
        print("ERROR: Could not access the Product Table! Shutting down!")
        Logger.Log("{}: {}".format(__name__, "ERROR: Could not access the Product Table! Shutting down!"))
        raise SystemExit

    allProducts = DataTypes.AllProducts()
    numberOfProducts = productSheet.cell(row = 1, column = 6).value

    print("-> Getting information for", numberOfProducts,"products:")
    Logger.Log("{}: {} {} {}".format(__name__, "-> Getting information for ", numberOfProducts," products:"))

    for it in range(0,numberOfProducts):
        allProducts.productId.insert(it, productSheet.cell(row = (2 + it), column = 1).value)
        allProducts.productName.insert(it, productSheet.cell(row = (2 + it), column = 2).value)
        allProducts.productPrice.insert(it, productSheet.cell(row = (2 + it), column = 3).value)

        print(allProducts.productId[it], allProducts.productName[it], allProducts.productPrice[it])
        Logger.Log("{}: {} {} {}".format(__name__, allProducts.productId[it], allProducts.productName[it], allProducts.productPrice[it]))

    return allProducts
